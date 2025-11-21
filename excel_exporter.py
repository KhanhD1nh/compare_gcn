from pathlib import Path
from typing import List, Dict
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment


def export_to_excel(results: List[Dict], output_path: Path):
    """
    Export results to Excel file. If file exists, append to it.
    If a file with the same name already exists in Excel, overwrite its prediction and result.
    
    Args:
        results: List of processing results
        output_path: Path to Excel output file
    """
    # Check if file already exists
    if output_path.exists():
        # Load existing workbook
        wb = openpyxl.load_workbook(output_path)
        ws = wb.active
        print(f"Loading existing file: {output_path}")
        
        # Build a map of existing files: {filename: row_index}
        existing_files = {}
        for row_idx in range(2, ws.max_row + 1):
            filename = ws.cell(row=row_idx, column=2).value
            if filename:
                existing_files[filename] = row_idx
    else:
        # Create new workbook
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "GCN Comparison Results"
        
        # Header
        headers = ["Số thứ tự", "Tên tệp GCN", "Dự đoán", "Kết quả"]
        ws.append(headers)
        
        # Format header
        header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF")
        for cell in ws[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center", vertical="center")
        
        existing_files = {}
    
    # Process results
    for result in results:
        pdf_filename = result["pdf_file"]
        
        # Check if this file already exists in the Excel
        if pdf_filename in existing_files:
            # Update existing row (overwrite prediction and result)
            row_idx = existing_files[pdf_filename]
            ws.cell(row=row_idx, column=3).value = result["predicted_gcn"]
            ws.cell(row=row_idx, column=4).value = result["comparison"]
            print(f"  Updated: {pdf_filename}")
        else:
            # Add new row
            new_index = ws.max_row  # This will be the index for the new row
            row = [
                new_index,
                pdf_filename,
                result["predicted_gcn"],
                result["comparison"]
            ]
            ws.append(row)
            existing_files[pdf_filename] = ws.max_row
            print(f"  Added: {pdf_filename}")
    
    # Format all data rows (from row 2 onwards)
    for row_idx in range(2, ws.max_row + 1):
        # Format result column (column 4)
        cell = ws.cell(row=row_idx, column=4)
        if cell.value == "Đúng":
            cell.fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
            cell.font = Font(color="006100", bold=True)
        elif cell.value == "Cần hiệu đính":
            cell.fill = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
            cell.font = Font(color="9C6500", bold=True)
        cell.alignment = Alignment(horizontal="center", vertical="center")
        
        # Center index column (column 1)
        ws.cell(row=row_idx, column=1).alignment = Alignment(horizontal="center", vertical="center")
    
    # Automatically adjust column width
    for column in ws.columns:
        max_length = 0
        column_letter = column[0].column_letter
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = min(max_length + 2, 50)
        ws.column_dimensions[column_letter].width = adjusted_width
    
    # Save file
    wb.save(output_path)
    print(f"\nSaved results to {output_path}")

