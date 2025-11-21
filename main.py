from pathlib import Path
from PIL import Image
import fitz
import time
import base64
import requests
import re
from typing import List, Dict, Optional, Tuple
from concurrent.futures import ThreadPoolExecutor, as_completed
from datetime import datetime
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment

from prompt import SYSTEM_INSTRUCTION_NUMBER_GCN
from config import Config


def find_all_gcn_pdfs(root_dir: Path) -> List[Path]:
    """
    Find all PDF files containing 'GCN' in the name (recursively)
    """
    gcn_files = sorted(root_dir.rglob('*GCN*.pdf'))
    print(f"Tìm thấy {len(gcn_files)} file GCN trong {root_dir}")
    return gcn_files


def extract_page2_to_base64(pdf_path: Path) -> Optional[str]:
    """
    Extract page 2 of PDF to base64 (without saving file)
    
    Args:
        pdf_path: Path to PDF file
        
    Returns:
        Base64 string of image, or None if error
    """
    try:
        with fitz.open(pdf_path) as doc:
            if doc.page_count < 2:
                return None
            
            # Get page 2 (index 1)
            page = doc.load_page(1)
            pix = page.get_pixmap(dpi=Config.RENDER_DPI)
            img = Image.frombytes('RGB', (pix.width, pix.height), pix.samples)
            gray_img = img.convert('L')
            
            # If image is horizontal, cut in half and keep right half
            if gray_img.width > gray_img.height:
                mid_x = gray_img.width // 2
                gray_img = gray_img.crop((mid_x, 0, gray_img.width, gray_img.height))
            
            # Convert to base64
            from io import BytesIO
            buffer = BytesIO()
            gray_img.save(buffer, format='PNG')
            img_bytes = buffer.getvalue()
            img_b64 = base64.b64encode(img_bytes).decode('utf8')
            
            return img_b64
            
    except Exception as e:
        print(f"✗ Error processing {pdf_path.name}: {e}")
        return None


def extract_gcn_with_llm(image_b64: str) -> Tuple[str, Optional[str]]:
    """
    Call LLM to extract GCN number from base64 image
    
    Args:
        image_b64: Base64 string of image
        
    Returns:
        Tuple of (GCN number extracted, error message if any)
    """
    try:
        payload = {
            "model": Config.MODEL,
            "messages": [
                {
                    "role": "system",
                    "content": [{"type": "text", "text": SYSTEM_INSTRUCTION_NUMBER_GCN}]
                },
                {
                    "role": "user",
                    "content": [
                        {"type": "text", "text": "Extract GCN number from image according to the specified format."},
                        {
                            "type": "image_url",
                            "image_url": {"url": f"data:image/png;base64,{image_b64}"}
                        }
                    ]
                }
            ],
            "temperature": Config.TEMPERATURE,
        }
        
        resp = requests.post(Config.LM_URL, json=payload, timeout=Config.API_TIMEOUT)
        resp.raise_for_status()
        data = resp.json()
        result = data["choices"][0]["message"]["content"]
        
        return result.strip(), None
        
    except requests.exceptions.Timeout:
        return "ERROR", "LLM API timeout"
    except requests.exceptions.RequestException as e:
        return "ERROR", f"LLM API error: {str(e)[:100]}"
    except Exception as e:
        return "ERROR", f"LLM error: {str(e)[:100]}"


def normalize_gcn_number(gcn_str: str) -> str:
    """
    Normalize GCN number: add space between letters and digits, uppercase
    
    Rules for prefix (bỏ ký tự đầu):
    - 4 letters → bỏ 2 đầu, giữ 2 cuối (SOAH → AH, SOCJ → CJ)
    - 3 letters → bỏ 2 đầu, giữ 1 cuối (SCV → V, SOB → B)
    - 2 letters → giữ nguyên (AA → AA)
    - 1 letter → giữ nguyên (D → D)
    
    Args:
        gcn_str: GCN number string to normalize
        
    Returns:
        Normalized GCN number (format: "XX 1234567")
    """
    # Remove special characters, only keep letters and numbers
    cleaned = re.sub(r'[^A-Za-z0-9]', '', gcn_str).upper()
    
    # Try to split letters and digits
    match = re.match(r'^([A-Z]+)(\d+)$', cleaned)
    if match:
        letters = match.group(1)
        digits = match.group(2)
        
        # Apply rules based on number of letters
        if len(letters) >= 4:
            # 4+ letters: bỏ 2 đầu, giữ 2 cuối
            letters = letters[-2:]
        elif len(letters) == 3:
            # 3 letters: bỏ 2 đầu, giữ 1 cuối
            letters = letters[-1:]
        # 1-2 letters: giữ nguyên
        
        return f"{letters} {digits}"
    
    return cleaned


def validate_filename_format(filename: str) -> Tuple[bool, str]:
    """
    Kiểm tra tên file có đúng quy định không
    
    Quy định:
    1. 1-2 ký tự đầu là chữ VIẾT HOA (A-Z)
    2. Sau đó là khoảng trắng (tùy chọn) và số
    3. Kết thúc bằng "-GCN.pdf"
    
    Args:
        filename: Tên file cần kiểm tra
        
    Returns:
        (is_valid, error_message)
    """
    # Pattern: [1-2 CHỮ HOA][space tùy chọn][SỐ]-GCN.pdf
    pattern = r'^([A-Z]{1,2})\s*(\d+)-GCN\.pdf$'
    
    match = re.match(pattern, filename)
    if not match:
        # Phân tích chi tiết lỗi
        if not filename.endswith('.pdf'):
            return False, "Không phải file PDF"
        
        if not filename.endswith('-GCN.pdf'):
            return False, "Không kết thúc bằng -GCN.pdf"
        
        # Check phần prefix
        prefix_match = re.match(r'^([a-zA-Z]{1,2})', filename)
        if not prefix_match:
            return False, "Không có chữ cái đầu tiên"
        
        prefix = prefix_match.group(1)
        if not prefix.isupper():
            return False, f"Chữ cái phải viết HOA (hiện tại: '{prefix}')"
        
        # Check có số không
        if not re.search(r'\d+', filename):
            return False, "Thiếu phần số"
        
        return False, "Sai format tên file"
    
    return True, ""


def extract_gcn_from_filename(filename: str) -> str:
    """
    Extract GCN number from filename
    
    Args:
        filename: Filename (example: "AA 01555158-GCN.pdf" hoặc "land-certificates-2025-11-15T04-10-51_AA 01555158-GCN.pdf")
        
    Returns:
        GCN number from filename (example: "AA 01555158")
    """
    # Find pattern: letter + number before "-GCN"
    # Pattern: [A-Z]{1,2}\s*\d+ (bất kỳ số chữ số nào)
    match = re.search(r'([A-Z]{1,2}\s*\d+)-GCN', filename, re.IGNORECASE)
    if match:
        gcn = match.group(1)
        return normalize_gcn_number(gcn)
    
    # Fallback: find different pattern
    match = re.search(r'_([A-Z]{1,2}\s*\d+)-', filename, re.IGNORECASE)
    if match:
        gcn = match.group(1)
        return normalize_gcn_number(gcn)
    
    return "UNKNOWN"


def compare_gcn(filename_gcn: str, predicted_gcn: str) -> str:
    """
    Compare GCN number from filename with AI prediction
    
    Args:
        filename_gcn: GCN number from filename
        predicted_gcn: GCN number predicted by AI
        
    Returns:
        "Đúng" or "Cần hiệu đính"
    """
    normalized_filename = normalize_gcn_number(filename_gcn)
    normalized_predicted = normalize_gcn_number(predicted_gcn)
    
    return "Đúng" if normalized_filename == normalized_predicted else "Cần hiệu đính"


def process_single_pdf(pdf_path: Path, index: int) -> Dict:
    """
    Process a single PDF file: extract page 2, call LLM, compare
    
    Args:
        pdf_path: Path to PDF file
        index: Index
        
    Returns:
        Dict containing processing result
    """
    start = time.time()
    result = {
        "index": index,
        "pdf_file": pdf_path.name,
        "pdf_path": str(pdf_path),
        "filename_gcn": "",
        "predicted_gcn": "",
        "comparison": "",
        "status": "pending",
        "error": None,
        "time": 0
    }
    
    try:
        # Step 0: Validate filename format
        is_valid, error_msg = validate_filename_format(pdf_path.name)
        if not is_valid:
            result["status"] = "skip"
            result["error"] = f"Sai tên file: {error_msg}"
            result["comparison"] = "N/A"
            return result
        
        # Step 1: Extract GCN number from filename
        filename_gcn = extract_gcn_from_filename(pdf_path.name)
        result["filename_gcn"] = filename_gcn
        
        # Step 2: Extract page 2 to base64
        img_b64 = extract_page2_to_base64(pdf_path)
        if not img_b64:
            result["status"] = "skip"
            result["error"] = "No page 2"
            result["comparison"] = "N/A"
            return result
        
        # Step 3: Call LLM to extract GCN number
        predicted_gcn_raw, llm_error = extract_gcn_with_llm(img_b64)
        
        # Normalize predicted GCN để hiển thị đúng format
        if predicted_gcn_raw != "ERROR":
            predicted_gcn = normalize_gcn_number(predicted_gcn_raw)
        else:
            predicted_gcn = predicted_gcn_raw
        
        result["predicted_gcn"] = predicted_gcn
        
        # Step 4: Compare
        if predicted_gcn != "ERROR" and filename_gcn != "UNKNOWN":
            comparison = compare_gcn(filename_gcn, predicted_gcn)
            result["comparison"] = comparison
            result["status"] = "success"
        else:
            result["comparison"] = "N/A"
            result["status"] = "error"
            # Chi tiết nguyên nhân lỗi
            if predicted_gcn == "ERROR" and filename_gcn == "UNKNOWN":
                result["error"] = f"Both LLM and filename failed. LLM: {llm_error}"
            elif predicted_gcn == "ERROR":
                result["error"] = f"LLM extraction failed: {llm_error}. Filename GCN: {filename_gcn}"
            elif filename_gcn == "UNKNOWN":
                result["error"] = f"Cannot parse GCN from filename. LLM predicted: {predicted_gcn}"
            else:
                result["error"] = "Cannot extract GCN number"
        
    except Exception as e:
        result["status"] = "error"
        result["error"] = str(e)
        result["comparison"] = "N/A"
    
    finally:
        result["time"] = time.time() - start
    
    return result


def process_batch_pdfs(
    pdf_files: List[Path],
    max_workers: int = 1
) -> List[Dict]:
    """
    Process batch of PDF files with multi-threading
    
    Args:
        pdf_files: List of PDF files to process
        max_workers: Number of parallel workers
        
    Returns:
        List of processing results
    """
    results = []
    
    with ThreadPoolExecutor(max_workers=max_workers) as executor:
        futures = {
            executor.submit(process_single_pdf, pdf, idx + 1): (pdf, idx + 1)
            for idx, pdf in enumerate(pdf_files)
        }
        
        for future in as_completed(futures):
            pdf_path, idx = futures[future]
            try:
                result = future.result()
                results.append(result)
                
                if result["status"] == "success":
                    status_icon = "✓" if result["comparison"] == "Đúng" else "⚠"
                    msg = f"{result['filename_gcn']} vs {result['predicted_gcn']} -> {result['comparison']}"
                elif result["status"] == "skip":
                    status_icon = "⚠"
                    msg = result.get('error', 'Skip')
                else:
                    status_icon = "✗"
                    msg = result.get('error', 'Error')
                
                print(f"{status_icon} [{result['time']:.2f}s] #{idx} {result['pdf_file']}: {msg}")
                
            except Exception as e:
                print(f"✗ {pdf_path.name}: Unexpected error - {e}")
    
    # Sort by index
    results.sort(key=lambda x: x["index"])
    return results


def export_to_excel(results: List[Dict], output_path: Path):
    """
    Export results to Excel file
    
    Args:
        results: List of processing results
        output_path: Path to Excel output file
    """
    # Create new workbook
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "GCN Comparison Results"
    
    # Header
    headers = ["Số thứ tự", "Tên tệp GCN", "Dự đoán", "Kết quả"]
    ws.append(headers)
    
    # Format header
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF")
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")
    
    # Add data
    for result in results:
        row = [
            result["index"],
            result["pdf_file"],
            result["predicted_gcn"],
            result["comparison"]
        ]
        ws.append(row)
    
    # Format result column
    for row_idx in range(2, len(results) + 2):
        cell = ws.cell(row=row_idx, column=4)
        if cell.value == "Đúng":
            cell.fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
            cell.font = Font(color="006100", bold=True)
        elif cell.value == "Cần hiệu đính":
            cell.fill = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
            cell.font = Font(color="9C6500", bold=True)
        cell.alignment = Alignment(horizontal="center", vertical="center")
    
    # Center index column
    for row_idx in range(2, len(results) + 2):
        ws.cell(row=row_idx, column=1).alignment = Alignment(horizontal="center", vertical="center")
    
    # Automatically adjust column width
    for column in ws.columns:
        max_length = 0
        column_letter = column[0].column_letter
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = min(max_length + 2, 50)
        ws.column_dimensions[column_letter].width = adjusted_width
    
    # Save file
    wb.save(output_path)
    print(f"\nSaved results to {output_path}")


def main():
    """Main function"""  
    start_total = time.time()
    
    # Step 1: Find all GCN files (recursively)
    print(f"\n[1/4] Searching for GCN files in {Config.INPUT_DIR}...")
    gcn_files = find_all_gcn_pdfs(Config.INPUT_DIR)
    
    if not gcn_files:
        print("No GCN files found. Exiting program.")
        return
    
    # Step 2: Ask user how many files to process
    print(f"\nFound {len(gcn_files)} GCN files.")
    try:
        user_input = input(f"Enter the number of files to process (1-{len(gcn_files)}, Enter = all): ").strip()
        if user_input:
            batch_size = int(user_input)
            batch_size = min(max(1, batch_size), len(gcn_files))
        else:
            batch_size = len(gcn_files)
    except:
        batch_size = len(gcn_files)
    
    selected_files = gcn_files[:batch_size]
    print(f"-> Will process {len(selected_files)} files")
    
    # Step 3: Process batch
    print(f"\n[2/4] Processing {len(selected_files)} files with {Config.MAX_WORKERS} workers...")
    results = process_batch_pdfs(
        pdf_files=selected_files,
        max_workers=Config.MAX_WORKERS
    )
    
    # Step 4: Statistics
    print("\n[3/4] Statistics:")
    success = sum(1 for r in results if r["status"] == "success")
    skip = sum(1 for r in results if r["status"] == "skip")
    error = sum(1 for r in results if r["status"] == "error")
    
    # Classify reasons for skip
    skip_bad_filename = sum(1 for r in results if r["status"] == "skip" and "Sai tên file" in str(r.get("error", "")))
    skip_no_page2 = sum(1 for r in results if r["status"] == "skip" and "No page 2" in str(r.get("error", "")))
    
    correct = sum(1 for r in results if r["comparison"] == "Đúng")
    incorrect = sum(1 for r in results if r["comparison"] == "Cần hiệu đính")
    
    print(f"  Success: {success}")
    print(f"  Skip: {skip}")
    if skip_bad_filename > 0:
        print(f"    - Bad filename: {skip_bad_filename}")
    if skip_no_page2 > 0:
        print(f"    - No page 2: {skip_no_page2}")
    print(f"  Error: {error}")
    print(f"\n  Comparison:")
    print(f"    - Correct: {correct}")
    print(f"    - Need correction: {incorrect}")
    if success > 0:
        accuracy = (correct / success) * 100
        print(f"    - Accuracy: {accuracy:.2f}%")
    
    # Step 5: Export Excel
    print("\n[4/4] Exporting results to Excel...")
    timestamp = datetime.now().strftime("%Y%m%d")
    excel_file = Path(f"gcn_comparison_{timestamp}.xlsx")
    export_to_excel(results, excel_file)
    
    # Summary
    total_time = time.time() - start_total
    print("\n" + "=" * 80)
    print(f"COMPLETE - Total time: {total_time:.2f} seconds")
    if len(results) > 0:
        print(f"Average: {total_time / len(results):.2f} seconds/file")
    print(f"Excel file: {excel_file.resolve()}")
    print("=" * 80)


if __name__ == "__main__":
    main()
